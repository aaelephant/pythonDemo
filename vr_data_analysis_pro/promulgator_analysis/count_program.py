#!/usr/local/bin/python
#-*- coding:utf-8 -*-

import sys
sys.path.append('../')

import sql_Module.connectMysql as conMysql
"""导入sql_Module文件夹下的模块"""
from collections import Counter
from openpyxl import Workbook
from openpyxl import load_workbook
import thread
# from openpyxl.compat import range
# from openpyxl.utils import get_column_letter


class CountProgram(object):
	"""docstring for CountProgram"""
	FILE_NAME = 'videoSid.xlsx'
	id_field = 'id'
	date_field = 'date'
	videoName_field = 'videoName'
	videoTags_field = 'videoTags'
	host_field = 'host'
	currentPageId_field = 'currentPageId'
	apkVersion_field = 'apkVersion'
	productModel_field = 'productModel'
	userId_field = 'userId'
	videoType_field = 'videoType'
	videoSid_field = 'videoSid'
	isMonocular_field = 'isMonocular'
	screenType_field = 'screenType'
	curBitType_field = 'curBitType'
	actionType_field = 'actionType'
	pv_field = 'pv'

	entry_name = 'count'
	entry_field = 'videoSid'

	def __init__(self):
		super(CountProgram, self).__init__()
		self.conms = conMysql.ConnectMysql()
		self.conms.configMdb()
		return
	def installFactors(self,factors):
		videoSid = ''
		startDate = ''
		endDate = ''
		actionType = ''
		if factors.has_key('videoSid'):
			print 'has_key:videoSid'
			videoSid = factors['videoSid']
		if factors.has_key('startDate'):
			print 'has_key:startDate'
			startDate = factors['startDate']
		if factors.has_key('endDate'):
			print 'has_key:endDate'
			endDate = factors['endDate']
		if factors.has_key('actionType'):
			print 'has_key:actionType'
			actionType = factors['actionType']
		tableName = self.conms.table_name

		sqls = []
		sql = ''
		andStr = ' and '
		if len(videoSid)>0:
			factor1 = "videoSid = '%s'" % (videoSid)
			sqls.append(factor1)
		if len(startDate):
			factor2 = "date >= '%s'" % (startDate)
			sqls.append(factor2)
		if len(endDate):
			factor3 =  "date <= '%s'" % (endDate)
			sqls.append(factor3)
		if len(actionType):
			factor4 = "actionType = '%s'" % (actionType)
		for factor in sqls:
			sql += (factor + andStr)
		if len(sql):
			sql = 'where ' + sql[0:len(sql)-len(andStr)]
		return sql
	def readWithFactor(self, pageNum, pageSize, factors):
		sql = self.installFactors(factors)
		dataTuples = self.conms.selectWithFactor("%s LIMIT %d, %d" %(sql, pageNum*pageSize,pageSize))
		tuples = self.conms.getColumns()

		column_names = []
		for cur in tuples:
			column_names.append(cur[0])
		# dataDics = []
		videoList = []
		for dataTuple in dataTuples:
			# itemDics = []
			for i in xrange(0,len(column_names)):
				# dataDic = {}
				# dataDic[column_names[i]] = dataTuple[i]
				# itemDics.append(dataDic)
				if column_names[i] == self.entry_field:
					videoList.append(dataTuple[i])
			# print 'itemDics:'+str(itemDics)+'\n'
			# dataDics.append(itemDics)
		r = Counter(videoList)
		print 'r:' + str(len(r.most_common()))
		# print 'videoList:'+str(len(videoList))+'条'
		return r
	def continueWriteToExcle(self, path):
		wb = load_workbook(path)
		print wb.get_sheet_names() 
	def writetoNewExcle(self,list):
		wb = Workbook()
		ws = wb.active
		ws.append([self.entry_field,self.entry_name])
		totalNum = len(list)
		for row in xrange(2, totalNum+2):
			index = row-2#totalNum+1-row#
			col = list[index][0]
			ws.cell(column=1, row=row, value=col)
			col = list[index][1]
			ws.cell(column=2, row=row, value=col)

		wb.save(self.FILE_NAME)
		self.conms.close()
		return
	
	def countPlayVideo(self, **factors):
		# startDate = '2016-12-15'
		# endDate = '2016-12-15'videoSid, startDate, endDate
		sql = self.installFactors(factors)
		count = self.conms.totalCount(sql)

		pageSize = 10000
		totalPageNum = count/pageSize
		print '总共有'+str(totalPageNum)+'页'
		# remainder = count%pageSize

		preCounter = self.readWithFactor(0, pageSize, factors)	
		for pageNum in xrange(1,totalPageNum):
			print '第'+str(pageNum)+'页'
			curCounter = self.readWithFactor(pageNum, pageSize, factors)	
			preCounter.update(curCounter)  
		curCounter = self.readWithFactor(totalPageNum, pageSize, factors)	
		preCounter.update(curCounter)
		list = preCounter.most_common()
		print 'list:'+str(len(list))
		self.writetoNewExcle(list)
		return self.FILE_NAME
	def mysqlColumns(self):
		tuples = self.conms.getColumns()
		return
	def main(self):
		list = self.countPlayVideo()
		self.writetoNewExcle(list)
		return
if __name__ == '__main__':
	
	# cp = CountProgram()
	# cp.continueWriteToExcle('continue.xlsx')
	# cp.mysqlColumns()
	cp.countPlayVideo()#(startDate = '2017-05-01',endDate = '2017-05-01', actionType='startplay')
	# cp.countPlayVideo('tvn8opvx1cno')