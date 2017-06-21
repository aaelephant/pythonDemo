#!/usr/local/bin/python
#-*- coding:utf-8 -*-

import pandas as pd
from openpyxl import Workbook

import sys
sys.path.append('../../../')

import sql_Module.connectMysql as conMysql
from sql_Module.mysql_const import const

class PlayStatisticsBase(object):
	"""docstring for PlayStatisticsBase"""
	# const.PLAY_COUNT = 'play_count'
	const.PLAY_USER_COUNT = 'user_count'
	def __init__(self):
		super(PlayStatisticsBase, self).__init__()
		self.conms = conMysql.ConnectMysql()
		self.conms.configMdb()

		self.loop = True
		self.chunkSize = 10000
	def writetoNewExcle(self, list, fileName):
		wb = Workbook()
		ws = wb.active
		ws.append([const.COLUMN_VIDEONAME,const.COLUMN_VIDEOSID,const.COLUMN_DATE,\
				   const.PLAY_COUNT,const.COLUMN_VIDEOTAGS,const.PLAY_USER_COUNT])
		totalNum = len(list)
		for row in xrange(2, totalNum+2):
			item = list[row-2]
			for column in xrange(1, len(item)+1):
				value = item[column-1]
				# if value is None:
					# value = ''
				ws.cell(column=column, row=row, value=value)	
			
			# col = list[index][1]
			# ws.cell(column=2, row=row, value=col)

		wb.save(fileName+'.xlsx')
		self.conms.close()
		return
	def getUserCount(self, videoSid):
		sql = 'SELECT '\
				  +const.COLUMN_USERID\
				  +' from '\
				  +const.TABLE_NAME\
				  +self.whereSql\
				  +' and '\
				  +const.COLUMN_VIDEOSID\
				  +const.EQUAL\
				  +" '"+videoSid+"' "
		print sql
		chunks = pd.read_sql(sql,\
	 			 con=self.conms.connector(), chunksize=self.chunkSize)
		df = pd.concat(chunks, ignore_index=False)
		r = df.groupby(const.COLUMN_USERID).size()
		count = len(r)
		print count
		return count
	def searchItemInfo(self, videoSid, playCount):
		sql = 'SELECT '\
				  +const.COLUMN_VIDEONAME\
				  +','\
				  +const.COLUMN_DATE\
				  +','\
				  +const.COLUMN_VIDEOTAGS\
				  +' from '\
				  +const.TABLE_NAME\
				  +self.whereSql\
				  +' and '\
				  +const.COLUMN_VIDEOSID\
				  +const.EQUAL\
				  +" '"+videoSid+"' "\
				  +' LIMIT 1'
		dataTuples = self.conms.select(sql)
		line = []
		if len(dataTuples):
		   line = dataTuples[0] 
		video = [line[0], videoSid, line[1],\
				playCount, line[2]
				]
		return video
	
	def installInfo(self, r):
		print len(r)
		num = 0
		dataList = []
		for index in r.index.values:
			video = self.searchItemInfo(index, r.at[index])
			userCount = self.getUserCount(index)
			video.append(userCount)
			dataList.append(video)
			num += 1
			if num>10:
				break
		return dataList

# ascending = True
		# if kwargs.has_key('ascending'):
		# 	ascending = kwargs['ascending']
		# columns = [const.COLUMN_VIDEOSID,const.COLUMN_USERID]
		# chunks = self.installChunks(columns, kwargs)
		# df = pd.concat(chunks, ignore_index=False)
		# # print type(df)
		# groupbyVideoSid = df.groupby(const.COLUMN_VIDEOSID)
		# print groupbyVideoSid.size()
		# groupbyuserId = df.groupby([const.COLUMN_VIDEOSID,const.COLUMN_USERID])
		# gr = groupbyuserId.size().to_frame()
		# r = gr.groupby(const.COLUMN_VIDEOSID)
		# # print r.size()
		# dfVideoSid = groupbyVideoSid.size().sort_values(ascending=False).to_frame()
		# dfUserId = r.size().to_frame()
		# print dfUserId
		# dfUserId.rename(columns={0: 'userCount'}, inplace=True)
		# df_videoPlayCount = self.videoPlayCountDataFrame(kwargs)
		# df_videoPlayUserCount = self.videoPlayUserCountDataFrame(kwargs)
		# dfMerge = pd.merge(df_videoPlayCount,df_videoPlayUserCount,left_index=True,right_index=True,how='left')
		# print dfMerge
		
		# dfMerge.to_excl("video_play_user.")
		# print len(groupbyVideoSid.size()["62a20fce53c44d6682f959b5df7ae31a"])
# df = df.groupby(by=['column_A'])['column_B'].sum()
# 　　生成的数据类型是Series,如果进一步需要将其转换为dataframe,可以调用Series中的to_frame()方法.
		# groupbyuserId = groupbyVideoSid.apply(const.COLUMN_USERID)
		# print groupbyuserId.size().index
		# groupbyVideoSidMergeUserId = pd.merge(groupbyVideoSid,groupbyuserId, key=const.COLUMN_VIDEOSID)

		# print groupbyVideoSidMergeUserId
		# dfGroupby.rename(columns={0: 'id'}, inplace=True);
		# print dfGroupby.index.values
		# columns1 = [const.COLUMN_VIDEOSID,const.COLUMN_VIDEONAME]
		# chunksInfo = self.installChunks(columns1,kwargs)
		# dfInfo = pd.concat(chunksInfo, ignore_index=False)
		# print dfInfo

		# writer = pd.ExcelWriter('live_play_statistics.xlsx')
		# dfMergeWithInfo.to_excel(writer,'Sheet1')
		# # df2.to_excel(writer,'Sheet2')
		# writer.save()
		# # for cur in dfMerge:
		# 	# print cur
		# return "live_play_statistics.xlsx"
		
		# dataList = self.installInfo(r)
		# self.writetoNewExcle(dataList,'live_play_statistics')
		# for level,subsetDF in r:
		# 	# print level
		# 	r2 = subsetDF.groupby(subsetDF[const.COLUMN_VIDEOSID]).size()
		# 	print r2.iloc[0]

			# break
		# r = r.sort_values(ascending=False)
		# for videoSid in r.index.values:
		
		# print r

		# dfMergeWithInfo = pd.merge(dfMerge,df_videoInfo,left_index=True,right_on=const.COLUMN_VIDEOSID,how='left').drop_duplicates()
		# dfMergeWithInfo.rename(columns={0: 'playCount'}, inplace=True)
		# dfMergeWithInfo = dfMergeWithInfo.reset_index()
		# del dfMergeWithInfo['index']